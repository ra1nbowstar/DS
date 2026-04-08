from services.finance_service import (
    FinanceService,
    cap_discounts_to_merchandise_total,
    max_coupon_total_yuan,
    parse_pending_coupon_ids,
)
from fastapi import APIRouter, HTTPException, Query, Request, Depends
from pydantic import BaseModel, Field, ConfigDict, AliasChoices, model_validator
from typing import Optional, List, Dict, Any, cast
from core.config import Settings, settings
from core.database import get_conn
from services.finance_service import split_order_funds
from core.config import VALID_PAY_WAYS, POINTS_DISCOUNT_RATE
from core.table_access import build_dynamic_select, get_table_structure, _quote_identifier
from decimal import Decimal, ROUND_DOWN
import uuid
from datetime import datetime, timedelta
from enum import Enum
import json
import threading
import time
from core.logging import get_logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any
from fastapi.responses import StreamingResponse

# ==================== 新增：导入 Redis 用于分布式锁 ====================
import redis
import redis.exceptions

logger = get_logger(__name__)
router = APIRouter()


# ==================== 新增：Redis 客户端初始化（带容错） ====================
def _get_redis_client():
    """获取 Redis 客户端，如果未配置则返回 None"""
    try:
        redis_host = getattr(settings, 'REDIS_HOST', 'localhost')
        redis_port = getattr(settings, 'REDIS_PORT', 6379)
        redis_db = getattr(settings, 'REDIS_DB', 0)

        client = redis.Redis(
            host=redis_host,
            port=redis_port,
            db=redis_db,
            decode_responses=True,
            socket_connect_timeout=2,
            socket_timeout=2
        )
        client.ping()
        return client
    except Exception as e:
        logger.warning(f"Redis 连接失败（将使用数据库兜底）: {e}")
        return None


# 全局 Redis 客户端
redis_client = _get_redis_client()


def _cancel_expire_orders():
    """每分钟扫描一次，把过期的 pending_pay 订单取消"""
    while True:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    now = datetime.now()
                    cur.execute("""
                        SELECT id, order_number
                        FROM orders
                        WHERE status='pending_pay'
                          AND expire_at IS NOT NULL
                          AND expire_at <= %s
                    """, (now,))
                    for o in cur.fetchall():
                        oid, ono = o["id"], o["order_number"]

                        # 删除该订单的待发放奖励记录
                        cur.execute(
                            "SELECT id FROM pending_rewards WHERE order_id = %s AND status = 'pending'",
                            (oid,)
                        )
                        rewards = cur.fetchall()
                        for reward in rewards:
                            cur.execute(
                                "DELETE FROM pending_rewards WHERE id = %s",
                                (reward['id'],)
                            )
                            print(f"[expire] 删除订单 {ono} 的待发放奖励记录: ID={reward['id']}")

                        # 回滚库存
                        cur.execute(
                            "SELECT product_id,quantity FROM order_items WHERE order_id=%s",
                            (oid,)
                        )
                        for it in cur.fetchall():
                            cur.execute(
                                "UPDATE product_skus SET stock=stock+%s WHERE product_id=%s",
                                (it["quantity"], it["product_id"])
                            )

                        # 改状态
                        cur.execute(
                            "UPDATE orders SET status='cancelled',updated_at=NOW() WHERE id=%s",
                            (oid,)
                        )
                        print(f"[expire] 订单 {ono} 已自动取消")
                    conn.commit()
        except Exception as e:
            print(f"[expire] error: {e}")
        time.sleep(60)


def start_order_expire_task():
    """由 api.order 包初始化时调用一次即可"""
    t = threading.Thread(target=_cancel_expire_orders, daemon=True)
    t.start()
    print("[expire] 订单过期守护线程已启动")


class OrderManager:
    @staticmethod
    def _build_orders_select(cursor) -> str:
        structure = get_table_structure(cursor, "orders")
        select_parts = []
        for field in structure['fields']:
            if field in structure['asset_fields']:
                select_parts.append(f"COALESCE({_quote_identifier(field)}, 0) AS {_quote_identifier(field)}")
            else:
                select_parts.append(_quote_identifier(field))
        return ", ".join(select_parts)

    @staticmethod
    def create(
            user_id: int,
            address_id: Optional[int],
            custom_addr: Optional[dict],
            specifications: Optional[str] = None,
            buy_now: bool = False,
            buy_now_items: Optional[List[Dict[str, Any]]] = None,
            delivery_way: str = "platform",
            points_to_use: Optional[Decimal] = None,
            coupon_id: Optional[int] = None,
            coupon_ids: Optional[List[int]] = None,
            idempotency_key: Optional[str] = None,
            merchant_id: Optional[int] = None
    ) -> Optional[str]:
        """创建订单（已增加幂等性校验，防止重复创建，支持多商家订单）"""
        lock_key = f"order:create:{user_id}"
        lock_acquired = False

        if redis_client:
            try:
                lock_acquired = redis_client.set(lock_key, idempotency_key or "1", nx=True, ex=5)
                if not lock_acquired:
                    logger.warning(f"用户 {user_id} 重复提交订单，Redis 锁拦截")
                    raise HTTPException(
                        status_code=429,
                        detail="订单创建中，请勿重复提交，或等待 5 秒后重试"
                    )
            except redis.exceptions.RedisError as e:
                logger.error(f"Redis 锁操作失败: {e}，将降级为数据库锁")
                lock_acquired = False

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:

                    if not buy_now:
                        cur.execute("""
                            SELECT order_number, status, created_at 
                            FROM orders 
                            WHERE user_id = %s 
                              AND created_at > DATE_SUB(NOW(), INTERVAL 1 MINUTE)
                              AND status != 'cancelled'
                            ORDER BY created_at DESC 
                            LIMIT 1
                        """, (user_id,))
                        recent_order = cur.fetchone()
                        if recent_order:
                            logger.warning(
                                f"用户 {user_id} 1 分钟内已有订单 {recent_order['order_number']}，拦截重复创建")
                            raise HTTPException(
                                status_code=400,
                                detail=f"您刚刚已创建订单 {recent_order['order_number']}，请勿重复提交"
                            )

                    if idempotency_key and redis_client:
                        used_key = f"order:idempotency:{idempotency_key}"
                        if redis_client.exists(used_key):
                            existing_order = redis_client.get(used_key)
                            logger.info(f"幂等 Key 重复，返回已存在订单: {existing_order}")
                            return existing_order

                    # ---------- 1. 组装订单明细 ----------
                    if buy_now:
                        if not buy_now_items:
                            raise HTTPException(status_code=422, detail="立即购买时 buy_now_items 不能为空")
                        items = []
                        product_merchant_ids = set()

                        for it in buy_now_items:
                            cur.execute("SELECT is_member_product, user_id, cash_only FROM products WHERE id = %s",
                                        (it["product_id"],))
                            prod = cur.fetchone()
                            if not prod:
                                raise HTTPException(status_code=404,
                                                    detail=f"products 表中不存在 id={it['product_id']}")

                            product_merchant_ids.add(prod.get("user_id") or 0)

                            sku_id = it.get("sku_id")
                            if not sku_id:
                                cur.execute("SELECT id FROM product_skus WHERE product_id = %s LIMIT 1",
                                            (it['product_id'],))
                                sku_row = cur.fetchone()
                                if sku_row:
                                    sku_id = sku_row.get('id')
                                else:
                                    raise HTTPException(status_code=422,
                                                        detail=f"商品 {it['product_id']} 无可用 SKU，请提供 sku_id")

                            if "price" not in it:
                                raise HTTPException(status_code=422,
                                                    detail=f"buy_now_items 必须包含 price 字段：product_id={it['product_id']}")

                            items.append({
                                "sku_id": sku_id,
                                "product_id": it["product_id"],
                                "quantity": it["quantity"],
                                "price": Decimal(str(it["price"])),
                                "is_vip": prod["is_member_product"],
                                "cash_only": prod["cash_only"]  # 新增
                            })

                        if len(product_merchant_ids) > 1:
                            raise HTTPException(
                                status_code=400,
                                detail="一笔订单只能包含同一商家的商品，请分开下单"
                            )

                        if merchant_id is None:
                            merchant_id = product_merchant_ids.pop() if product_merchant_ids else 0
                    else:
                        cur.execute("""
                            SELECT c.product_id,
                                c.sku_id,
                                c.quantity,
                                s.price,
                                p.is_member_product AS is_vip,
                                p.user_id as merchant_id,
                                c.specifications,
                                p.cash_only
                            FROM cart c
                            JOIN product_skus s ON s.id = c.sku_id
                            JOIN products p ON p.id = c.product_id
                            WHERE c.user_id = %s AND c.selected = 1
                        """, (user_id,))
                        items = cur.fetchall()
                        if not items:
                            return None

                        merchant_ids = set(item.get("merchant_id") or 0 for item in items)
                        if len(merchant_ids) > 1:
                            raise HTTPException(
                                status_code=400,
                                detail="购物车中包含不同商家的商品，请分开结算"
                            )

                        if merchant_id is None:
                            merchant_id = merchant_ids.pop() if merchant_ids else 0

                    merchant_id = int(merchant_id or 0)

                    # ---------- 2. 优惠券商品类型验证（支持多张券叠加） ----------
                    has_vip = any(i["is_vip"] for i in items)
                    coupon_id_list: List[int] = []
                    if coupon_ids:
                        coupon_id_list = sorted({int(x) for x in coupon_ids})
                    elif coupon_id is not None:
                        coupon_id_list = [int(coupon_id)]

                    sum_coupons = Decimal('0')

                    # 先计算商品总额，可用于优惠券/积分验证
                    total = sum(Decimal(str(i["quantity"])) * Decimal(str(i["price"])) for i in items)

                    for cid in coupon_id_list:
                        cur.execute("""
                            SELECT id, amount, applicable_product_type, status, valid_from, valid_to, user_id
                            FROM coupons 
                            WHERE id = %s AND user_id = %s AND status = 'unused'
                            FOR UPDATE
                        """, (cid, user_id))
                        coupon = cur.fetchone()

                        if not coupon:
                            raise HTTPException(status_code=400, detail=f"优惠券不存在、已被使用或不属于当前用户: id={cid}")

                        today = datetime.now().date()
                        if not (coupon['valid_from'] <= today <= coupon['valid_to']):
                            raise HTTPException(status_code=400, detail="优惠券不在有效期内")

                        applicable_type = coupon['applicable_product_type']
                        if applicable_type == 'member_only' and not has_vip:
                            raise HTTPException(status_code=400, detail="该优惠券仅限会员商品使用")
                        if applicable_type == 'normal_only' and has_vip:
                            raise HTTPException(status_code=400, detail="该优惠券仅限普通商品使用")

                        sum_coupons += Decimal(str(coupon['amount']))

                    # 检查是否有 cash_only 商品
                    has_cash_only = any(i.get("cash_only") for i in items)
                    if has_cash_only:
                        if points_to_use and points_to_use > 0:
                            raise HTTPException(status_code=400, detail="该商品只能用现金支付，不能使用积分")
                        if coupon_id_list:
                            raise HTTPException(status_code=400, detail="该商品只能用现金支付，不能使用优惠券")

                    # 积分先封顶，再校验券叠加上限 = ceil(原价 − 积分抵扣) 到元
                    pt_use = points_to_use or Decimal('0')
                    pd_raw = pt_use * POINTS_DISCOUNT_RATE
                    points_discount = min(pd_raw, total)
                    if POINTS_DISCOUNT_RATE and POINTS_DISCOUNT_RATE > 0:
                        pt_use = (points_discount / POINTS_DISCOUNT_RATE).quantize(
                            Decimal('0.0001'), rounding=ROUND_DOWN
                        )
                    else:
                        pt_use = Decimal('0')

                    max_c = max_coupon_total_yuan(total, points_discount)
                    if sum_coupons > max_c:
                        raise HTTPException(
                            status_code=400,
                            detail=f"优惠券叠加面额不能超过{max_c}元（商品金额扣减积分抵扣后，向上取整到元）"
                        )

                    coupon_discount = sum_coupons
                    points_to_use = pt_use
                    # 与结算侧一致的安全收斂（边界数值）
                    coupon_discount, points_discount, pt_use = cap_discounts_to_merchandise_total(
                        total, coupon_discount, pt_use
                    )
                    points_to_use = pt_use
                    final_amount = total - points_discount - coupon_discount

                    pending_coupon_ids_json = json.dumps(coupon_id_list) if coupon_id_list else None
                    pending_coupon_single = coupon_id_list[0] if len(coupon_id_list) == 1 else None

                    # 判断是否零元订单（新增）
                    is_zero_order = final_amount <= Decimal('0')

                    # 判断初始状态和过期时间
                    if is_zero_order:
                        # 与支付回调一致：快递/平台配送 → 待发货；自提 → 待收货（结算成功后会再 update_status，含全虚拟→已完成）
                        init_status = "pending_recv" if delivery_way == "pickup" else "pending_ship"
                        expire_at = None
                    else:
                        init_status = "pending_pay"
                        expire_at = datetime.now() + timedelta(hours=12)

                    # ---------- 3. 地址信息 ----------
                    if delivery_way == "pickup":
                        consignee_name = consignee_phone = province = city = district = shipping_address = ""
                    elif custom_addr:
                        consignee_name = custom_addr.get("consignee_name")
                        consignee_phone = custom_addr.get("consignee_phone")
                        province = custom_addr.get("province", "")
                        city = custom_addr.get("city", "")
                        district = custom_addr.get("district", "")
                        shipping_address = custom_addr.get("detail", "")
                    else:
                        raise HTTPException(status_code=422, detail="必须上传收货地址或选择自提")

                    # ---------- 4. 订单主表 ----------
                    order_number = (
                            datetime.now().strftime("%Y%m%d%H%M%S") +
                            str(user_id) +
                            uuid.uuid4().hex[:16]
                    )

                    cur.execute("""
                        INSERT INTO orders(
                            user_id, merchant_id, order_number, total_amount, original_amount, status, is_vip_item,
                            consignee_name, consignee_phone,
                            province, city, district, shipping_address, delivery_way,
                            pay_way, auto_recv_time, refund_reason, expire_at,
                            pending_points, pending_coupon_id, pending_coupon_ids,
                            points_discount, coupon_discount)
                        VALUES (%s, %s, %s, %s, %s, %s, %s,
                                %s, %s, %s, %s, %s, %s, %s,
                                'wechat', %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        user_id, merchant_id, order_number, final_amount, total, init_status, has_vip,
                        consignee_name, consignee_phone,
                        province, city, district, shipping_address, delivery_way,
                        datetime.now() + timedelta(days=3),
                        specifications,
                        expire_at,
                        points_to_use or Decimal('0'),
                        pending_coupon_single,
                        pending_coupon_ids_json,
                        points_discount,
                        coupon_discount
                    ))
                    oid = cur.lastrowid

                    # ---------- 5. 库存校验 & 扣减 ----------
                    structure = get_table_structure(cur, "product_skus")
                    has_stock_field = 'stock' in structure['fields']
                    stock_select = (
                        f"COALESCE({_quote_identifier('stock')}, 0) AS {_quote_identifier('stock')}"
                        if has_stock_field and 'stock' in structure['asset_fields']
                        else _quote_identifier('stock')
                    ) if has_stock_field else "0 AS stock"

                    for i in items:
                        cur.execute(f"SELECT {stock_select} FROM {_quote_identifier('product_skus')} WHERE id=%s",
                                    (i['sku_id'],))
                        result = cur.fetchone()
                        current_stock = result.get('stock', 0) if result else 0
                        if current_stock < i["quantity"]:
                            raise HTTPException(status_code=400, detail=f"SKU {i['sku_id']} 库存不足")

                    # ---------- 6. 写订单明细 ----------
                    for i in items:
                        cur.execute("""
                            INSERT INTO order_items(order_id, product_id, sku_id, quantity, unit_price, total_price)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            oid, i["product_id"], i["sku_id"], i["quantity"],
                            i["price"], Decimal(str(i["quantity"])) * Decimal(str(i["price"]))
                        ))

                    # ---------- 7. 扣库存 ----------
                    if has_stock_field:
                        for i in items:
                            cur.execute("UPDATE product_skus SET stock = stock - %s WHERE id = %s",
                                        (i["quantity"], i['sku_id']))

                    # ---------- 8. 清空购物车（仅购物车结算场景） ----------
                    if not buy_now:
                        cur.execute("DELETE FROM cart WHERE user_id = %s AND selected = 1", (user_id,))

                    # 结算逻辑：零元或普通订单均在 finance_service 内处理状态和优惠券
                    if is_zero_order:
                        from services.finance_service import FinanceService
                        fs = FinanceService()
                        fs.settle_order(
                            order_no=order_number,
                            user_id=user_id,
                            order_id=oid,
                            points_to_use=points_to_use or Decimal('0'),
                            coupon_discount=coupon_discount,
                            external_conn=conn
                        )

                    if idempotency_key and redis_client:
                        used_key = f"order:idempotency:{idempotency_key}"
                        redis_client.setex(used_key, 86400, order_number)

                    conn.commit()
                    logger.info(f"订单创建成功: {order_number}, 用户: {user_id}, 商家: {merchant_id}")

                    return {
                        "order_number": order_number,
                        "need_pay": not is_zero_order
                    }

        finally:
            if lock_acquired and redis_client:
                try:
                    redis_client.delete(lock_key)
                except Exception as e:
                    logger.error(f"释放 Redis 锁失败: {e}")

    @staticmethod
    def list_by_user(user_id: int, status: Optional[str] = None):
        """按用户查询订单列表，附带首件商品和规格字段。"""
        with get_conn() as conn:
            with conn.cursor() as cur:
                select_fields = OrderManager._build_orders_select(cur)
                sql = f"SELECT {select_fields} FROM orders WHERE user_id = %s"
                params = [user_id]
                if status:
                    sql += " AND status = %s"
                    params.append(status)
                sql += " ORDER BY created_at DESC"
                cur.execute(sql, tuple(params))
                orders = cur.fetchall()

                for o in orders:
                    cur.execute(
                        """
                        SELECT oi.*, p.name
                        FROM order_items oi
                        JOIN products p ON oi.product_id = p.id
                        WHERE oi.order_id = %s
                        LIMIT 1
                        """,
                        (o["id"],)
                    )
                    first_item = cur.fetchone()
                    o["first_product"] = first_item
                    o["specifications"] = o.get("refund_reason")

                return orders

    @staticmethod
    def list_by_merchant(
            merchant_id: int,
            status: Optional[str] = None,
            start_date: Optional[str] = None,
            end_date: Optional[str] = None,
            page: int = 1,
            page_size: int = 20
    ) -> Dict[str, Any]:
        """按商家查询订单列表（支持分页、状态筛选、时间范围筛选）"""
        with get_conn() as conn:
            with conn.cursor() as cur:
                where_conditions = ["merchant_id = %s"]
                params = [merchant_id]

                if status:
                    where_conditions.append("status = %s")
                    params.append(status)

                if start_date:
                    where_conditions.append("DATE(created_at) >= %s")
                    params.append(start_date)

                if end_date:
                    where_conditions.append("DATE(created_at) <= %s")
                    params.append(end_date)

                where_clause = " AND ".join(where_conditions)

                count_sql = f"SELECT COUNT(*) as total FROM orders WHERE {where_clause}"
                cur.execute(count_sql, tuple(params))
                total = cur.fetchone()["total"]

                amount_sql = f"""
                    SELECT 
                        COALESCE(SUM(total_amount), 0) as total_amount,
                        COALESCE(SUM(CASE WHEN status = 'completed' THEN total_amount ELSE 0 END), 0) as completed_amount
                    FROM orders 
                    WHERE {where_clause}
                """
                cur.execute(amount_sql, tuple(params))
                amount_stats = cur.fetchone()

                select_fields = OrderManager._build_orders_select(cur)
                offset = (page - 1) * page_size
                sql = f"""
                    SELECT {select_fields} 
                    FROM orders 
                    WHERE {where_clause}
                    ORDER BY created_at DESC
                    LIMIT %s OFFSET %s
                """
                query_params = params + [page_size, offset]
                cur.execute(sql, tuple(query_params))
                orders = cur.fetchall()

                for o in orders:
                    cur.execute(
                        """
                        SELECT oi.*, p.name as product_name, p.cover as product_cover
                        FROM order_items oi
                        JOIN products p ON oi.product_id = p.id
                        WHERE oi.order_id = %s
                        """,
                        (o["id"],)
                    )
                    o["items"] = cur.fetchall()

                    cur.execute(
                        "SELECT id, name, mobile, avatar FROM users WHERE id = %s",
                        (o["user_id"],)
                    )
                    user_info = cur.fetchone()
                    o["user_info"] = user_info

                return {
                    "list": orders,
                    "pagination": {
                        "page": page,
                        "page_size": page_size,
                        "total": total,
                        "total_pages": (total + page_size - 1) // page_size
                    },
                    "statistics": {
                        "total_amount": float(amount_stats["total_amount"]),
                        "completed_amount": float(amount_stats["completed_amount"]),
                        "order_count": total
                    }
                }

    @staticmethod
    def detail(order_number: str) -> Optional[dict]:
        """查询单个订单详情（含用户、地址、商品明细、商家信息）。"""
        with get_conn() as conn:
            with conn.cursor() as cur:
                select_fields = OrderManager._build_orders_select(cur)
                cur.execute(
                    f"SELECT {select_fields} FROM orders WHERE order_number=%s LIMIT 1",
                    (order_number,)
                )
                order = cur.fetchone()
                if not order:
                    return None

                order_id = order.get("id")
                user_id = order.get("user_id")
                merchant_id = order.get("merchant_id") or 0

                cur.execute(
                    """
                    SELECT oi.*, p.name AS product_name, p.is_member_product, p.cover AS product_cover, p.cash_only
                    FROM order_items oi
                    LEFT JOIN products p ON oi.product_id = p.id
                    WHERE oi.order_id = %s
                    """,
                    (order_id,)
                )
                items = cur.fetchall()

                user_info = None
                if user_id:
                    cur.execute(
                        "SELECT id, name, mobile, avatar, member_level, member_points FROM users WHERE id=%s",
                        (user_id,)
                    )
                    user_info = cur.fetchone()

                merchant_info = None
                if merchant_id and merchant_id > 0:
                    cur.execute(
                        """
                        SELECT u.id, u.name, u.mobile, u.avatar, u.wechat_sub_mchid,
                               ms.store_name, ms.store_logo_image_id, ms.store_address
                        FROM users u
                        LEFT JOIN merchant_stores ms ON ms.user_id = u.id
                        WHERE u.id = %s AND u.is_merchant = 1
                        """,
                        (merchant_id,)
                    )
                    merchant_info = cur.fetchone()

                # ===== 待支付优惠券（支持多张叠加面额） =====
                pending_coupon_ids = parse_pending_coupon_ids(order)
                pending_coupon_amount = None
                if pending_coupon_ids:
                    ph = ",".join(["%s"] * len(pending_coupon_ids))
                    cur.execute(
                        f"SELECT COALESCE(SUM(amount), 0) AS s FROM coupons WHERE id IN ({ph})",
                        pending_coupon_ids,
                    )
                    srow = cur.fetchone()
                    if srow and srow.get("s") is not None:
                        pending_coupon_amount = float(srow["s"])
                pending_coupon_id = order.get("pending_coupon_id")
                # ======================================

                address = {
                    "consignee_name": order.get("consignee_name"),
                    "consignee_phone": order.get("consignee_phone"),
                    "province": order.get("province"),
                    "city": order.get("city"),
                    "district": order.get("district"),
                    "detail": order.get("shipping_address"),
                }

                return {
                    "order_info": order,
                    "user": user_info,
                    "merchant": merchant_info,
                    "address": address,
                    "items": items,
                    "specifications": order.get("refund_reason"),
                    # ===== 新增返回 pending 字段 =====
                    "pending_points": float(order.get("pending_points") or 0),
                    "pending_coupon_id": pending_coupon_id,
                    "pending_coupon_ids": pending_coupon_ids,
                    "pending_coupon_amount": pending_coupon_amount,
                    # ===== 方便前端直接读取已抵扣金额 =====
                    "points_discount": float(order.get("points_discount") or 0),
                    "coupon_discount": float(order.get("coupon_discount") or 0),
                }

    @staticmethod
    def update_status(order_number: str, new_status: str, reason: Optional[str] = None,
                      external_conn=None) -> bool:
        """统一的订单状态更新，支持外部连接复用。"""

        def _apply_update(cur) -> bool:
            cur.execute("SHOW COLUMNS FROM orders")
            cols = {row.get("Field") for row in cur.fetchall()}

            updates = ["status=%s", "updated_at=NOW()"]
            params: List[Any] = [new_status]

            if reason:
                for col in ("status_reason", "remark"):
                    if col in cols:
                        updates.append(f"{col}=%s")
                        params.append(reason)
                        break

            if new_status in ("pending_ship", "pending_recv") and "paid_at" in cols:
                updates.append("paid_at=COALESCE(paid_at, NOW())")
            if new_status == "pending_recv" and "shipped_at" in cols:
                updates.append("shipped_at=COALESCE(shipped_at, NOW())")
            if new_status == "completed" and "completed_at" in cols:
                updates.append("completed_at=COALESCE(completed_at, NOW())")

            params.append(order_number)
            cur.execute(f"UPDATE orders SET {', '.join(updates)} WHERE order_number=%s", tuple(params))
            return cur.rowcount > 0

        if external_conn:
            cur = external_conn.cursor()
            try:
                return _apply_update(cur)
            finally:
                cur.close()
        else:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    updated = _apply_update(cur)
                    conn.commit()
                    return updated

    @staticmethod
    def confirm_receive(order_number: str, user_id: Optional[int] = None) -> Dict[str, Any]:
        """
        用户确认收货（仅更新本地订单状态，不再主动同步微信）
        """
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT id, user_id, status, order_number, transaction_id, delivery_way
                       FROM orders WHERE order_number=%s""",
                    (order_number,)
                )
                order = cur.fetchone()

                if not order:
                    return {"ok": False, "message": "订单不存在"}

                if user_id and order['user_id'] != user_id:
                    return {"ok": False, "message": "无权操作该订单"}

                if order['status'] != 'pending_recv':
                    return {"ok": False, "message": f"订单状态不正确，当前状态：{order['status']}"}

                cur.execute(
                    "UPDATE orders SET status='completed', completed_at=NOW() WHERE id=%s",
                    (order['id'],)
                )
                conn.commit()

        logger.info(f"用户 {user_id or order['user_id']} 确认收货成功，订单号：{order_number}")

        # 不再主动调用微信接口，微信侧订单状态将由用户通过官方确认收货组件触发的回调更新
        return {"ok": True, "message": "确认收货成功"}

    @staticmethod
    def export_to_excel(order_numbers: List[str]) -> bytes:
        """
        导出订单详情（包含资金拆分明细）
        生成两个工作表：订单详情、资金拆分明细
        """
        account_type_map = {
            "merchant_balance": "商家余额",
            "public_welfare": "公益基金",
            "maintain_pool": "平台维护",
            "subsidy_pool": "周补贴池",
            "director_pool": "联创奖励",
            "shop_pool": "社区店",
            "city_pool": "城市运营中心",
            "branch_pool": "大区分公司",
            "fund_pool": "事业发展基金",
            "company_points": "公司积分账户",
            "company_balance": "公司余额账户",
            "platform_revenue_pool": "平台收入池（会员商品）",
            "wx_applyment_fee": "微信进件手续费",
            "income": "收入",
            "expense": "支出"
        }

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "订单详情"

        headers1 = [
            "订单号", "商家ID", "商家名称", "订单状态", "总金额", "原始金额", "积分抵扣", "实付金额",
            "支付方式", "配送方式", "是否会员订单",
            "用户ID", "用户姓名", "用户手机号",
            "收货人", "收货电话", "省份", "城市", "区县", "详细地址",
            "商品信息", "商品规格", "下单时间", "支付时间", "发货时间"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws2 = wb.create_sheet(title="资金拆分")
        headers2 = [
            "订单号", "商家ID", "账户类型", "变动金额", "变动后余额",
            "流水类型", "备注", "创建时间"
        ]

        for col_idx, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        row_idx1 = 2
        row_idx2 = 2

        with get_conn() as conn:
            with conn.cursor() as cur:
                for order_number in order_numbers:
                    order_data = OrderManager.detail(order_number)
                    if not order_data:
                        continue

                    order_info = order_data["order_info"]
                    user_info = order_data["user"]
                    merchant_info = order_data.get("merchant")
                    address = order_data["address"] or {}
                    items = order_data["items"]
                    specifications = order_data.get("specifications") or {}

                    total = Decimal(str(order_info.get("total_amount", 0)))
                    points_discount = Decimal(str(order_info.get("points_discount", 0)))
                    actual_pay = total - points_discount

                    product_info = "\n".join([
                        f"{item.get('product_name', '')} x{item.get('quantity', 0)} @¥{item.get('unit_price', 0)}"
                        for item in items
                    ])

                    spec_str = ""
                    if isinstance(specifications, dict):
                        spec_str = "\n".join([f"{k}: {v}" for k, v in specifications.items()])

                    shipped_at = order_info.get("shipped_at", "")
                    if order_info.get("delivery_way") == "pickup":
                        shipped_at = order_info.get("paid_at", "")

                    row_data1 = [
                        order_info.get("order_number", ""),
                        order_info.get("merchant_id", 0),
                        merchant_info.get("store_name") or merchant_info.get("name",
                                                                             "平台自营") if merchant_info else "平台自营",
                        order_info.get("status", ""),
                        float(total),
                        float(order_info.get("original_amount", 0)),
                        float(points_discount),
                        float(actual_pay),
                        order_info.get("pay_way", "wechat"),
                        order_info.get("delivery_way", "platform"),
                        "是" if order_info.get("is_member_order") else "否",
                        user_info.get("id", ""),
                        user_info.get("name", ""),
                        user_info.get("mobile", ""),
                        address.get("consignee_name", ""),
                        address.get("consignee_phone", ""),
                        address.get("province", ""),
                        address.get("city", ""),
                        address.get("district", ""),
                        address.get("detail", ""),
                        product_info,
                        spec_str,
                        order_info.get("created_at", ""),
                        order_info.get("paid_at", ""),
                        shipped_at
                    ]

                    for col_idx, value in enumerate(row_data1, 1):
                        cell = ws1.cell(row=row_idx1, column=col_idx, value=value)
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
                        cell.border = thin_border
                        if col_idx in [5, 6, 7, 8]:
                            cell.number_format = '¥#,##0.00'

                    row_idx1 += 1

                    cur.execute("""
                        SELECT account_type, change_amount, balance_after, 
                               flow_type, remark, created_at
                        FROM account_flow 
                        WHERE remark LIKE %s
                        ORDER BY created_at ASC
                    """, (f"%{order_number}%",))

                    flows = cur.fetchall()

                    platform_fee = float(actual_pay) * 0.2

                    for flow in flows:
                        account_type_en = flow.get("account_type", "")

                        if account_type_en == "merchant_balance":
                            display_amount = f"{int(platform_fee)}雨点"
                            account_type_cn = "商家余额"
                            balance_after_display = "-"
                            is_platform_fee_row = True
                        else:
                            account_type_cn = account_type_map.get(account_type_en, account_type_en)
                            display_amount = float(flow.get("change_amount", 0))
                            balance_after_display = float(flow.get("balance_after", 0))
                            is_platform_fee_row = False

                        row_data2 = [
                            order_number,
                            order_info.get("merchant_id", 0),
                            account_type_cn,
                            display_amount,
                            balance_after_display,
                            flow.get("flow_type", ""),
                            flow.get("remark", ""),
                            flow.get("created_at", "")
                        ]

                        for col_idx, value in enumerate(row_data2, 1):
                            cell = ws2.cell(row=row_idx2, column=col_idx, value=value)
                            cell.alignment = Alignment(vertical="center")
                            cell.border = thin_border
                            if col_idx in [4, 5] and not is_platform_fee_row:
                                cell.number_format = '¥#,##0.00'

                        row_idx2 += 1

        for column in ws1.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws1.column_dimensions[column_letter].width = min(max_length + 2, 50)

        for column in ws2.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws2.column_dimensions[column_letter].width = min(max_length + 2, 50)

        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        return excel_data.getvalue()


# ---------- 新增：获取 FinanceService 实例 ----------
def get_finance_service() -> FinanceService:
    """获取 FinanceService 实例（直接返回）"""
    return FinanceService()


# ---------------- 请求模型 ----------------
class DeliveryWay(str, Enum):
    platform = "platform"
    pickup = "pickup"


class OrderCreate(BaseModel):
    """同时接受 snake_case 与小程序常用 camelCase 请求体字段。"""

    model_config = ConfigDict(populate_by_name=True)

    @model_validator(mode="before")
    @classmethod
    def normalize_coupon_payload(cls, data: Any) -> Any:
        """兼容嵌套 body、单值券 ID、字符串列表等小程序常见传参。"""
        if not isinstance(data, dict):
            return data
        d: Dict[str, Any] = dict(data)
        # 一层嵌套：{ "data": { ... } } / payload / order
        for wrap in ("data", "payload", "order"):
            inner = d.get(wrap)
            if isinstance(inner, dict):
                rest = {k: v for k, v in d.items() if k != wrap}
                d = {**inner, **rest}
                break

        def _coerce_id_list(v: Any) -> Optional[List[int]]:
            if v is None:
                return None
            if isinstance(v, bool):
                return None
            if isinstance(v, int):
                return [v]
            if isinstance(v, float):
                return [int(v)]
            if isinstance(v, str):
                s = v.strip()
                if not s:
                    return None
                if s.startswith("[") and s.endswith("]"):
                    try:
                        parsed = json.loads(s)
                        if isinstance(parsed, list):
                            return [int(x) for x in parsed if x is not None]
                    except (ValueError, TypeError, json.JSONDecodeError):
                        pass
                out: List[int] = []
                for part in s.replace("，", ",").split(","):
                    part = part.strip()
                    if part.isdigit():
                        out.append(int(part))
                return out or None
            if isinstance(v, list):
                out = []
                for x in v:
                    if x is None:
                        continue
                    if isinstance(x, dict) and "id" in x:
                        out.append(int(x["id"]))
                    else:
                        out.append(int(x))
                return out or None
            return None

        # 统一写到 snake_case，后续 Field 仍可通过别名覆盖
        if d.get("coupon_ids") is None and d.get("couponIds") is not None:
            d["coupon_ids"] = d.pop("couponIds")
        ids = _coerce_id_list(d.get("coupon_ids"))
        if ids is not None:
            d["coupon_ids"] = ids

        if d.get("coupon_id") is None and d.get("couponId") is not None:
            d["coupon_id"] = d.pop("couponId")
        # 其它常见键名
        if not d.get("coupon_ids") and d.get("coupon_id") is None:
            for alt in ("selectedCouponId", "selected_coupon_id", "couponID"):
                if alt not in d or d[alt] is None or d[alt] == "":
                    continue
                try:
                    d["coupon_ids"] = [int(d.pop(alt))]
                    break
                except (TypeError, ValueError):
                    continue
        raw_objs = d.get("coupons") or d.get("couponList") or d.get("coupon_list")
        if raw_objs and not d.get("coupon_ids") and d.get("coupon_id") is None:
            if isinstance(raw_objs, list):
                idl: List[int] = []
                for it in raw_objs:
                    if isinstance(it, dict) and it.get("id") is not None:
                        idl.append(int(it["id"]))
                if idl:
                    d["coupon_ids"] = idl

        cid = d.get("coupon_id")
        if cid is not None and cid != "" and not isinstance(cid, bool):
            try:
                d["coupon_id"] = int(cid)
            except (TypeError, ValueError):
                d["coupon_id"] = None

        return d

    user_id: int = Field(validation_alias=AliasChoices("user_id", "userId"))
    delivery_way: DeliveryWay = Field(
        default=DeliveryWay.platform,
        validation_alias=AliasChoices("delivery_way", "deliveryWay"),
    )
    address_id: Optional[int] = Field(
        default=None,
        validation_alias=AliasChoices("address_id", "addressId"),
    )
    custom_address: Optional[dict] = Field(
        default=None,
        validation_alias=AliasChoices("custom_address", "customAddress"),
    )
    specifications: Optional[str] = None
    buy_now: bool = Field(default=False, validation_alias=AliasChoices("buy_now", "buyNow"))
    buy_now_items: Optional[List[Dict[str, Any]]] = Field(
        default=None,
        validation_alias=AliasChoices("buy_now_items", "buyNowItems"),
    )
    points_to_use: Optional[Decimal] = Field(
        default=Decimal('0'),
        validation_alias=AliasChoices("points_to_use", "pointsToUse"),
    )
    coupon_id: Optional[int] = Field(
        default=None,
        validation_alias=AliasChoices("coupon_id", "couponId"),
    )
    coupon_ids: Optional[List[int]] = Field(
        default=None,
        validation_alias=AliasChoices("coupon_ids", "couponIds"),
    )
    idempotency_key: Optional[str] = Field(
        default=None,
        validation_alias=AliasChoices("idempotency_key", "idempotencyKey"),
    )
    merchant_id: Optional[int] = Field(
        default=None,
        validation_alias=AliasChoices("merchant_id", "merchantId"),
    )


class OrderPay(BaseModel):
    order_number: str
    pay_way: str
    coupon_id: Optional[int] = None
    coupon_ids: Optional[List[int]] = None
    points_to_use: Optional[Decimal] = Decimal('0')


class StatusUpdate(BaseModel):
    order_number: str
    new_status: str
    reason: Optional[str] = None


class WechatPayParams(BaseModel):
    appId: str
    timeStamp: str
    nonceStr: str
    package: str
    signType: str
    paySign: str


class ConfirmReceiveRequest(BaseModel):
    order_number: str
    # 已废弃：不再使用微信确认结果，仅保留订单号用于本地状态更新


class MerchantOrdersQuery(BaseModel):
    status: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    page: int = 1
    page_size: int = 20


# ---------------- 路由 ----------------
@router.post("/create", summary="创建订单")
def create_order(body: OrderCreate):
    logger.info(
        "创建订单入参(券): user_id=%s buy_now=%s coupon_id=%s coupon_ids=%s",
        body.user_id,
        body.buy_now,
        body.coupon_id,
        body.coupon_ids,
    )
    result = OrderManager.create(
        body.user_id,
        body.address_id,
        body.custom_address,
        specifications=body.specifications,
        buy_now=body.buy_now,
        buy_now_items=body.buy_now_items,
        delivery_way=body.delivery_way,
        points_to_use=body.points_to_use,
        coupon_id=body.coupon_id,
        coupon_ids=body.coupon_ids,
        idempotency_key=body.idempotency_key,
        merchant_id=body.merchant_id
    )
    if not result or not result.get("order_number"):
        raise HTTPException(status_code=422, detail="购物车为空或地址缺失")
    return {
        "order_number": result["order_number"],
        "need_pay": result.get("need_pay", True)
    }


# 兼容旧客户端：/order/pay -> 代理到 /wechat-pay/create-order
@router.post("/pay", summary="兼容：创建支付参数 (旧路径)")
async def pay_compat(request: Request):
    try:
        # 延迟导入以避免循环依赖
        from api.wechat_pay import routes as wechat_routes

        return await wechat_routes.create_jsapi_order(request)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{user_id}", summary="查询用户订单列表")
def list_orders(user_id: int, status: Optional[str] = None):
    return OrderManager.list_by_user(user_id, status)


@router.get("/detail/{order_number}", summary="查询订单详情")
def order_detail(order_number: str):
    d = OrderManager.detail(order_number)
    if not d:
        raise HTTPException(status_code=404, detail="订单不存在")
    return d


@router.post("/status", summary="更新订单状态")
def update_status(body: StatusUpdate):
    return {"ok": OrderManager.update_status(body.order_number, body.new_status, body.reason)}


@router.get("/merchant/{merchant_id}", summary="查询商家订单列表")
def list_merchant_orders(
        merchant_id: int,
        status: Optional[str] = Query(None, description="订单状态筛选"),
        start_date: Optional[str] = Query(None, description="开始日期(YYYY-MM-DD)"),
        end_date: Optional[str] = Query(None, description="结束日期(YYYY-MM-DD)"),
        page: int = Query(1, ge=1, description="页码"),
        page_size: int = Query(20, ge=1, le=100, description="每页数量")
):
    result = OrderManager.list_by_merchant(
        merchant_id=merchant_id,
        status=status,
        start_date=start_date,
        end_date=end_date,
        page=page,
        page_size=page_size
    )
    return result


# 修改后
@router.post("/confirm-receive", summary="用户确认收货")
def confirm_receive(body: ConfirmReceiveRequest):
    """
    用户确认收货接口（已移除微信订单管理同步）

    仅更新本地订单状态，不再与微信发货管理系统交互
    """
    result = OrderManager.confirm_receive(body.order_number)
    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result["message"])
    return result


def auto_receive_task(db_cfg: dict = None):
    """自动收货守护进程（不再发放积分）"""
    import threading
    import time
    from datetime import datetime

    def run():
        while True:
            try:
                from core.database import get_conn
                with get_conn() as conn:
                    with conn.cursor() as cur:
                        now = datetime.now()
                        cur.execute(
                            "SELECT id, order_number, total_amount FROM orders "
                            "WHERE status='pending_recv' AND auto_recv_time<=%s",
                            (now,)
                        )
                        for row in cur.fetchall():
                            order_id = row["id"]
                            order_number = row["order_number"]

                            cur.execute(
                                "UPDATE orders SET status='completed' WHERE id=%s",
                                (order_id,)
                            )
                            conn.commit()
                            logger.debug(f"[auto_receive] 订单 {order_number} 已自动完成。")
            except Exception as e:
                logger.error(f"[auto_receive] 异常: {e}")
            time.sleep(3600)

    t = threading.Thread(target=run, daemon=True)
    t.start()
    logger.info("自动收货守护进程已启动（不再发放积分）")


class OrderExportRequest(BaseModel):
    order_numbers: List[str]


class OrderExportByTimeRequest(BaseModel):
    start_time: str
    end_time: str
    status: Optional[str] = None

# ---------- 新增请求模型 ----------
class DailySummaryExportRequest(BaseModel):
    start_date: str = Field(..., description="开始日期 yyyy-MM-dd")
    end_date: str = Field(..., description="结束日期 yyyy-MM-dd")
    include_detail: bool = Field(True, description="是否包含明细（优惠券/积分流水）")


@router.post("/export", summary="导出订单详情到Excel")
def export_orders(body: OrderExportRequest):
    if not body.order_numbers:
        raise HTTPException(status_code=422, detail="订单号列表不能为空")
    if len(body.order_numbers) > 1000:
        raise HTTPException(status_code=422, detail="单次导出订单数不能超过1000个")
    try:
        excel_data = OrderManager.export_to_excel(body.order_numbers)
        return StreamingResponse(
            BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=orders_export.xlsx"}
        )
    except Exception as e:
        logger.error(f"导出订单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/export/by-time", summary="按时间范围导出订单")
def export_orders_by_time(body: OrderExportByTimeRequest):
    try:
        start = datetime.strptime(body.start_time, "%Y-%m-%d %H:%M:%S")
        end = datetime.strptime(body.end_time, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        raise HTTPException(
            status_code=422,
            detail="时间格式错误，请使用：YYYY-MM-DD HH:MM:SS"
        )
    if end < start:
        raise HTTPException(
            status_code=422,
            detail="结束时间不能早于开始时间"
        )
    if (end - start).days > 31:
        raise HTTPException(
            status_code=422,
            detail="时间范围不能超过31天"
        )
    with get_conn() as conn:
        with conn.cursor() as cur:
            if body.status:
                sql = """
                    SELECT order_number 
                    FROM orders 
                    WHERE created_at >= %s 
                      AND created_at <= %s 
                      AND status = %s 
                    ORDER BY created_at DESC 
                    LIMIT 500
                """
                cur.execute(sql, (body.start_time, body.end_time, body.status))
            else:
                sql = """
                    SELECT order_number 
                    FROM orders 
                    WHERE created_at >= %s 
                      AND created_at <= %s 
                    ORDER BY created_at DESC 
                    LIMIT 500
                """
                cur.execute(sql, (body.start_time, body.end_time))
            rows = cur.fetchall()
            order_numbers = [row["order_number"] for row in rows]
    if not order_numbers:
        raise HTTPException(
            status_code=404,
            detail="该时间段内没有符合条件的订单"
        )
    try:
        excel_data = OrderManager.export_to_excel(order_numbers)
        start_str = body.start_time[:10].replace("-", "")
        end_str = body.end_time[:10].replace("-", "")
        filename = f"orders_{start_str}_to_{end_str}.xlsx"
        if body.status:
            filename = f"orders_{body.status}_{start_str}_to_{end_str}.xlsx"
        return StreamingResponse(
            BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"按时间导出订单失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")

# ---------- 新增：日报表/月报表导出 ----------
# 将 await 改为直接调用
@router.post("/export/daily-summary", summary="导出日报表/月报表")
async def export_daily_summary(
    request: DailySummaryExportRequest,
    service: FinanceService = Depends(get_finance_service)
):
    try:
        excel_data = service.export_daily_summary(   # ← 移除 await
            start_date=request.start_date,
            end_date=request.end_date,
            include_detail=request.include_detail
        )
        filename = f"daily_summary_{request.start_date}_to_{request.end_date}.xlsx"
        return StreamingResponse(
            BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"导出日报表失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

start_order_expire_task()
# start_wechat_status_sync_task()  # 已废弃